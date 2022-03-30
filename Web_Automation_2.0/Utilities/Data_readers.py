import pdb
import sys
from Utilities.logger_log import customLogger
import openpyxl

log = customLogger()


class Data_readers:

    def __init__(self):
        self.filename = "/home/alethea/PycharmProjects/Web_Automation_2.0/Utilities/Xpath's.xlsx"
        self.sheet = ''
        self.excel = None

    def excel_file_object(self):
        try:
            log.info("Loading the workbook")
            self.excel = openpyxl.load_workbook(self.filename)
            return self.excel.active
        except:
            log.error(sys.exc_info())

    def bring_locator(self, *page_name):
        xpath_holder = []
        try:
            # pdb.set_trace()
            for name in page_name:
                xpaths = {}
                global usr_index
                self.sheet = self.excel_file_object()
                if not self.sheet.max_column:
                    raise Exception("There is no Columns in the sheet")
                col_count = self.sheet.max_column
                row_count = self.sheet.max_row
                col_list = [self.sheet.cell(row=1, column=i).value for i in range(1, col_count + 1)]
                assert col_count == 4, "Column should not be more than 3 --> name should be {} {}  {}".format(
                    col_list[0], col_list[1], col_list[2], col_list[3])
                for row in range(2, row_count + 1):
                    cell_val = self.sheet.cell(row=row, column=2)
                    if str(cell_val.value) == name:
                        usr_index = self.getting_index(row_count, page_name=name)
                        break
                for row in usr_index:
                    obj_name = self.sheet.cell(row + 1, column=3)
                    obj_locator = self.sheet.cell(row + 1, column=4)
                    xpaths[obj_name.value] = obj_locator.value
                xpath_holder.append(xpaths)
            return xpath_holder
        except:
            log.error("error in bring locator {}".format(sys.exc_info()))
        finally:
            del xpath_holder

    def getting_index(self, rowcount, page_name):
        cell_index = []
        try:
            if self.sheet.cell(row=1, column=1).value == 'S.No':
                for x in range(2, rowcount + 1):
                    cell_val = self.sheet.cell(row=x, column=2)
                    if str(cell_val.value) == page_name:
                        user_index = self.sheet.cell(row=x, column=1)
                        cell_index.append(user_index.value)
                return cell_index
            else:
                raise Exception("S.No Column does't exists")
        except:
            log.error("getting index is failed {}".format(sys.exc_info()))

    def bring_control(self,controls,sheetname='Controls'):
        global get_index
        controls_sender,control_store ={},[]
        try:
            excel_sheet=openpyxl.load_workbook(self.filename)
            controls_sheet = excel_sheet['{}'.format(sheetname)]
            controls_sheet_status= excel_sheet.active
            if not controls_sheet:
                log.error("Sheet is not activated {}".format(sys.exc_info()))
            row_count = controls_sheet.max_row
            col_count = controls_sheet.max_column
            for row in range(2,row_count+1):
                cel_value = controls_sheet.cell(row=row,column=2).value
                if str(cel_value)== controls:
                    get_index = self.getting_index(rowcount=row_count,page_name=sheetname)
                    break
                for index in get_index:
                    sub_controls = controls_sheet.cell(row=index,column=col_count+2).value
                    main_controls = controls_sheet.cell(row=index,column=col_count+3).value
                    controls_sender[sub_controls]=main_controls
                control_store.append(controls_sender)
            return control_store
        except:
            log.error("bring_control failed {}".format(sys.exc_info()))



    def data_readers_teardown(self):
        try:
            self.excel.close()
        except:
            log.error("can't able to close the file")



if __name__ == '__main__':
    s1 = Data_readers()
    print(s1.bring_control(controls='Text'))
